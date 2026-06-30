#!/usr/bin/env python3
"""
Extract chapter stats + bracket data from Excel files (or Google Sheets CSVs).

USAGE — from Excel (one-time setup):
  python3 extract.py --source excel --excel-dir /path/to/excel/files

USAGE — from Google Sheets (run by GitHub Actions nightly):
  python3 extract.py --source sheets

Google Sheets config: edit SHEETS_CONFIG below with your published sheet IDs.
To get a sheet ID: open your Google Sheet, copy the long ID from the URL.
  https://docs.google.com/spreadsheets/d/YOUR_SHEET_ID_HERE/edit
To get the GID (tab ID): click the tab, look at the URL: ...#gid=123456789
"""

import json
import os
import sys
import argparse
import csv
import io
from pathlib import Path

try:
    import openpyxl
except ImportError:
    pass

try:
    import urllib.request
except ImportError:
    pass

# ── CONFIG ────────────────────────────────────────────────────────────────────
# Direct "Publish to web" CSV URLs from Google Sheets.
# To update these: open your Google Sheet → File → Share → Publish to web
# → select the Chapter Stats tab → CSV → Copy Link → paste below.
SHEETS_CONFIG = {
    2025: "https://docs.google.com/spreadsheets/d/e/2PACX-1vTubxnna3F_XctN7E0HVDw5gtXw5-qOGeiPEK_UgWS22F4dewQKW5USDPLOLWdXiP9RB1sGdl8x35H3/pub?gid=1374929055&single=true&output=csv",
    2024: "https://docs.google.com/spreadsheets/d/e/2PACX-1vRZqobY63U93Jwl4y4Y7UMJEnGUh7dEMEYrSFVc5B5BmQEWQ0CKAXNOlBbmBTNfnP4eDe078q37mIPy/pub?gid=689869648&single=true&output=csv",
    2023: "https://docs.google.com/spreadsheets/d/e/2PACX-1vSzJQHMYkLFfal6c9ScR4kVg4pW8wP2ZElYwjsF58W1D-G9fR0OMHMjI3CSazBjbcAAdYxAs4dFd0PZ/pub?gid=1317846843&single=true&output=csv",
    2022: "https://docs.google.com/spreadsheets/d/e/2PACX-1vSLeMi-qPJxxgBV3yWHKJKc0vYfZsansYxN2FnHP1l0Ust_tBZNw_RPpQXRxHIMOrQM37VD0KNFqIVW/pub?gid=1731092132&single=true&output=csv",
}

EXCEL_FILES = {
    2022: "Texas_Playoff_Brackets_2022.xlsx",
    2023: "Texas_Playoff_Brackets_2023_Final.xlsx",
    2024: "Texas_Playoff_Brackets_2024_LFG.xlsx",
    2025: "2025_Texas_Football_Officials_Data_Final.xlsx",
}

CHAPTER_STATS_SHEET = {
    2022: "Chapter Stats ",
    2023: "Chapter Stats ",
    2024: "Chapter Stats ",
    2025: "2025 Chapter Stats ",
}

CLASSIFICATIONS = ["1AD1","1AD2","2AD1","2AD2","3AD1","3AD2","4AD1","4AD2","5AD1","5AD2","6AD1","6AD2"]

BRACKET_SHEETS = ["1A D1","1A D2","2A D1","2A D2","3A D1","3A D2","4A D1","4A D2","5A D1","5A D2","6A D1","6A D2"]

ROUND_NAMES = {
    1: "Area",
    2: "Regionals",
    3: "Quarterfinals",
    4: "Semifinals",
    5: "State Championship",
    6: "State Championship",
}

# Known TASO chapter codes (3-letter abbreviations)
KNOWN_CHAPTERS = {
    "ABI","AMA","AUS","BEA","STX","COM","CSC","CTX","DAL","ELP","FTW","HOU",
    "NTX","PBC","PVC","RGV","SAT","SFA","SAN","SPC","TYL","WAC","WACO",
}

OUT_DIR = Path(__file__).parent.parent / "public" / "data"


def safe_int(v):
    try:
        f = float(v)
        return int(f)
    except (TypeError, ValueError):
        return 0


def parse_chapter_stats_rows(rows):
    """
    Parse Chapter Stats sheet rows into structured JSON.

    The sheet has 3 section-blocks, each covering 2 weeks:
      Block 1: Weeks 1 & 2
      Block 2: Weeks 3 & 4
      Block 3: Weeks 5 & 6 (Championship)

    Each block starts with a 'Total Games' header row, followed by
    chapter data rows (one row per chapter), then metadata rows.
    """
    chapters = {}

    # Find all block-start row indices.
    # Normally col[0]=='Total Games'; in 2024 col[0] is None but col[1]=='1AD1'.
    block_starts = []
    for i, row in enumerate(rows):
        if not row:
            continue
        col0 = str(row[0] or "").strip()
        col1 = str(row[1] or "").strip() if len(row) > 1 else ""
        if col0 == "Total Games" or col1 == "1AD1":
            block_starts.append(i)

    week_pairs = [(1, 2), (3, 4), (5, 6)]

    for block_idx, start_idx in enumerate(block_starts[:3]):
        if block_idx >= len(week_pairs):
            break
        w1, w2 = week_pairs[block_idx]

        # Determine end of this block: next block_start or end of sheet
        end_idx = block_starts[block_idx + 1] if block_idx + 1 < len(block_starts) else len(rows)

        # Iterate only the rows within THIS block
        for row in rows[start_idx + 1 : end_idx]:
            if not row or not row[0]:
                continue
            code = str(row[0]).strip()

            # Skip metadata rows
            if code in ("Missing:", "% complete:", "average complete", "Total Games"):
                continue
            if len(code) > 6:
                continue
            # Must look like a chapter code (2-5 uppercase letters/digits)
            if not code.replace("0","").replace("1","").replace("2","").replace("3","").replace("4","").replace("5","").replace("6","").replace("7","").replace("8","").replace("9","").isalpha():
                continue

            # Try to get full chapter name (usually appears around col 30)
            name = ""
            for ci in range(29, min(36, len(row))):
                v = str(row[ci] or "").strip()
                if v and len(v) > 2 and not v.replace(".","").replace("%","").replace("-","").isnumeric():
                    name = v.strip()
                    break

            if code not in chapters:
                chapters[code] = {"code": code, "name": name or code, "weeks": {}}
            elif name and chapters[code]["name"] == code:
                chapters[code]["name"] = name

            def parse_week(col_start):
                clf = {}
                for ci, cls in enumerate(CLASSIFICATIONS):
                    idx = col_start + ci
                    clf[cls] = safe_int(row[idx] if idx < len(row) else 0)
                total_idx = col_start + 12
                cum_idx = col_start + 13
                total = safe_int(row[total_idx] if total_idx < len(row) else 0)
                cumulative = safe_int(row[cum_idx] if cum_idx < len(row) else 0)
                return {"classifications": clf, "total": total, "cumulative": cumulative}

            chapters[code]["weeks"][str(w1)] = parse_week(1)
            chapters[code]["weeks"][str(w2)] = parse_week(16)

    # Set final total from the highest cumulative value seen across all weeks
    for code, ch in chapters.items():
        max_cum = max((w["cumulative"] for w in ch["weeks"].values()), default=0)
        ch["total"] = max_cum

    return list(chapters.values())


def parse_bracket_sheet(ws_rows, classification, year):
    """Parse a single bracket sheet into a list of game records."""
    games = []

    if len(ws_rows) < 3:
        return games

    # Detect if there's a Seeding column (added in 2023+)
    header_row = ws_rows[1] if len(ws_rows) > 1 else []
    has_seeding = any("seed" in str(c or "").lower() for c in header_row[:6])

    # Column offsets for team/score pairs per week
    # Without seeding: week1_team=col2, week1_score=col3, week2=col4/5 ...
    # With seeding:    seeding=col2, week1_team=col3, week1_score=col4 ...
    if has_seeding:
        week_cols = [(3, 4), (5, 6), (7, 8), (9, 10), (11, 12)]
    else:
        week_cols = [(2, 3), (4, 5), (6, 7), (8, 9), (10, 11)]

    def to_score(v):
        try:
            f = float(v)
            return None if f == 0.01 else int(f)  # 0.01 = forfeit placeholder
        except (TypeError, ValueError):
            return None

    def is_venue_string(s):
        """Return True if this looks like a venue/location note, not a team name."""
        if not s:
            return True
        keywords = ["@", "Fri", "Sat", "Sun", "Mon", "AT&T", "Stadium", "pm", "am", "Dec", "Nov"]
        return any(k in s for k in keywords) or len(s) > 55

    # Rows come in pairs: top team (row_a) + bottom team (row_b)
    i = 2  # skip 2 header rows
    current_chapter = {}  # week_num -> chapter code

    while i < len(ws_rows) - 1:
        row_a = ws_rows[i]
        row_b = ws_rows[i + 1] if i + 1 < len(ws_rows) else None

        # Update the running chapter from col 1 (chapter col) in row_a
        if row_a and row_a[1] and str(row_a[1]).strip():
            raw = str(row_a[1]).strip()
            # Only treat it as a chapter code if it looks right
            if 2 <= len(raw) <= 5 and raw.upper() == raw:
                current_chapter[1] = raw  # week 1 chapter from this pair

        # Try to find chapter codes embedded in week columns (for later weeks)
        # In the brackets, chapter for week N sometimes appears in row_a's score col
        if row_a:
            for wk_num, (tc, sc) in enumerate(week_cols[1:], start=2):
                if sc < len(row_a):
                    val = str(row_a[sc] or "").strip()
                    if 2 <= len(val) <= 5 and val.upper() == val and val.isalpha():
                        current_chapter[wk_num] = val

        if not row_a or not any(row_a):
            i += 2
            continue

        for week_num, (team_col, score_col) in enumerate(week_cols, start=1):
            if team_col >= len(row_a):
                continue

            team1 = str(row_a[team_col] or "").strip()
            score1_raw = row_a[score_col] if score_col < len(row_a) else None
            team2 = str(row_b[team_col] or "").strip() if row_b and team_col < len(row_b) else ""
            score2_raw = row_b[score_col] if row_b and score_col < len(row_b) else None

            if is_venue_string(team1) or is_venue_string(team2):
                continue
            if not team1 or not team2:
                continue

            score1 = to_score(score1_raw)
            score2 = to_score(score2_raw)

            winner = None
            if score1 is not None and score2 is not None:
                winner = team1 if score1 > score2 else team2

            chapter = current_chapter.get(week_num, "")

            games.append({
                "year": year,
                "classification": classification,
                "week": week_num,
                "round": ROUND_NAMES.get(week_num, f"Week {week_num}"),
                "chapter": chapter,
                "team1": team1,
                "score1": score1,
                "team2": team2,
                "score2": score2,
                "winner": winner,
            })

        i += 2

    return games


def extract_from_excel(excel_path, year):
    """Extract data from a single Excel file."""
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    stats_sheet_name = CHAPTER_STATS_SHEET.get(year, "Chapter Stats ")
    chapters = []
    if stats_sheet_name in wb.sheetnames:
        ws = wb[stats_sheet_name]
        rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
        chapters = parse_chapter_stats_rows(rows)

    games = []
    for sheet_name in BRACKET_SHEETS:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
            games.extend(parse_bracket_sheet(rows, sheet_name, year))

    return {"year": year, "chapters": chapters, "games": games}


def fetch_sheet_csv(url):
    """Fetch a published Google Sheet CSV URL."""
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        return resp.read().decode("utf-8")


def csv_to_rows(csv_text):
    """Parse CSV text into a list of tuples (with numeric coercion)."""
    def coerce(v):
        if v == "":
            return None
        try:
            return int(v)
        except ValueError:
            try:
                return float(v)
            except ValueError:
                return v
    reader = csv.reader(io.StringIO(csv_text))
    return [tuple(coerce(c) for c in row) for row in reader]


def extract_from_sheets(year, url):
    """Extract data from a published Google Sheets CSV URL."""
    print(f"  Fetching {year} chapter stats from Google Sheets...")
    csv_text = fetch_sheet_csv(url)
    rows = csv_to_rows(csv_text)
    chapters = parse_chapter_stats_rows(rows)
    return {"year": year, "chapters": chapters, "games": []}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", choices=["excel", "sheets"], default="sheets")
    parser.add_argument("--excel-dir", default=".", help="Directory containing Excel files")
    args = parser.parse_args()

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    all_years = []

    for year in [2022, 2023, 2024, 2025]:
        print(f"Processing {year}...")
        if args.source == "excel":
            excel_dir = Path(args.excel_dir)
            excel_file = excel_dir / EXCEL_FILES[year]
            if not excel_file.exists():
                print(f"  WARNING: {excel_file} not found, skipping")
                continue
            data = extract_from_excel(excel_file, year)
        else:
            url = SHEETS_CONFIG.get(year)
            if not url:
                print(f"  SKIP: Google Sheets URL not configured for {year}")
                continue
            data = extract_from_sheets(year, url)

        out_file = OUT_DIR / f"{year}.json"
        with open(out_file, "w") as f:
            json.dump(data, f, indent=2)
        print(f"  → wrote {out_file} ({len(data['chapters'])} chapters, {len(data['games'])} games)")
        all_years.append(year)

    import datetime
    index = {"years": all_years, "updated": datetime.date.today().isoformat()}
    with open(OUT_DIR / "index.json", "w") as f:
        json.dump(index, f, indent=2)
    print("Done.")


if __name__ == "__main__":
    main()
